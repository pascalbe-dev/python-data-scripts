#!/usr/bin/env python3

from argparse import ArgumentParser
from openpyxl import load_workbook
import warnings


def calculateVariableDifference(relevantVariableCells, dataRow, variableName, firstTimeSpot, lastTimeSpot):
    firstHeaderCell = list(filter(
        lambda x: str.startswith(x.value, firstTimeSpot + "_") and variableName in x.value, relevantVariableCells))[0]
    lastHeaderCell = list(filter(
        lambda x: str.startswith(x.value, lastTimeSpot + "_") and variableName in x.value, relevantVariableCells))[0]
    firstDataCellValue = list(dataRow)[firstHeaderCell.column-1].value
    lastDataCellValue = list(dataRow)[lastHeaderCell.column-1].value
    absolutDifference = lastDataCellValue - firstDataCellValue
    relativeDifference = absolutDifference / firstDataCellValue

    # print('##########################################################')
    # print('Printing information about the calculated cell differences')
    # print('Variable Name:', variableName)
    # print('Start TimeSpot:', firstTimeSpot)
    # print('End TimeSpot:', lastTimeSpot)
    # print('Value of the start data cell:', firstDataCellValue)
    # print('Value of the last data cell:', lastDataCellValue)
    # print('Absolute difference', absolutDifference)
    # print('Relative difference', relativeDifference, 'In percentage:',
    #       str(round(relativeDifference * 100, 2)) + "%")
    # print('##########################################################')
    return (absolutDifference, relativeDifference)


# arguments of the script
parser = ArgumentParser()
parser.add_argument("-f", "--file", dest="filename",
                    help="file to read", metavar="FILENAME", default="testfile.xlsx")
parser.add_argument("-s", "--sheetname", dest="sheetname",
                    help="sheet to read", metavar="SHEETNAME")
args = parser.parse_args()
sheetname = args.sheetname
filename = args.filename
timespots = ['Pre', 'Mid', 'Post']

# get excel sheet
workbook = load_workbook(filename)
worksheet = workbook.active if sheetname == None else workbook[sheetname]

# get required data
rowList = list(worksheet.rows)

# get all variables
# get all items in the top row
allVariableCells = list(rowList[0])
# remove empty variables
notEmptyVariableCells = list(
    filter(lambda x: x.value != None, allVariableCells))
# get variables for all required time spots (--> maybe there are additional variables which we do not care about)
relevantVariableCells = list(filter(lambda x: any(str.startswith(
    x.value, timespot + "_") for timespot in timespots), notEmptyVariableCells))
# get distinct variable names (without time spot)
allVariableNames = list(
    map(lambda x: x.value.partition('_')[-1], relevantVariableCells))
distinctVariableNames = list(dict.fromkeys(allVariableNames))
# remove variable names which do not have the appropriate amount of values (--> like amount of time spots)
relevantVariableNames = list(filter(
    lambda x: allVariableNames.count(x) == len(timespots), distinctVariableNames))
# catch variable names which are not equal for all time spots
irrelevantVariableNames = [
    x for x in allVariableNames if x not in relevantVariableNames]
irrelevantVariableNames.sort()
warnings.warn('The following variables have different names amongst the different time spots:\n\n' +
              str(irrelevantVariableNames))

# go over all relevant variables and calculate differences between first and last timespot
# then write those differences to the excel file
numberOfUsedColumns = len(allVariableCells)
firstTimeSpot = timespots[0]
lastTimeSpot = timespots[2]
for (variableIndex, variableName) in enumerate(relevantVariableNames):
    # name the headers of the columns appropriately
    firstColumnNumber = numberOfUsedColumns + 1 + (variableIndex * 2)
    absoluteHeaderCell = worksheet.cell(row=1, column=firstColumnNumber)
    relativeHeaderCell = worksheet.cell(row=1, column=firstColumnNumber + 1)
    absoluteHeaderCell.value = "Absolute-" + firstTimeSpot + \
        "-" + lastTimeSpot + "-" + variableName
    relativeHeaderCell.value = "Relative-" + firstTimeSpot + \
        "-" + lastTimeSpot + "-" + variableName
    absoluteHeaderCell.style = 'Headline 3'
    relativeHeaderCell.style = 'Headline 3'

    # go over all rows, calculate differences between values for the specific variable
    # and finally write the difference values to extra columns, whose headers were already written
    for (rowIndex, row) in enumerate(rowList[1:]):
        difference = calculateVariableDifference(
            relevantVariableCells, row, variableName, firstTimeSpot, lastTimeSpot)
        absoluteDataCell = worksheet.cell(
            row=rowIndex + 2, column=firstColumnNumber)
        relativeDataCell = worksheet.cell(
            row=rowIndex + 2, column=firstColumnNumber + 1)
        absoluteDataCell.value = difference[0]
        relativeDataCell.value = difference[1]
        absoluteDataCell.style = '20 % - Accent1'
        relativeDataCell.style = '20 % - Accent2'


workbook.save(filename + "-timeChanges.xlsx")
